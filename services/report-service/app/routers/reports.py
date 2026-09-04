from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
import io
from datetime import datetime, timedelta

from ..database import get_db
from ..models import Meeting, MeetingParticipant, AgendaItem, Report, User
from ..schemas import ReportOut, AttendanceSummary, AttendanceReportItem
from ..dependencies import get_current_user, AnyAuthenticated
from ..reports_data import gather_meeting_data
from ..generators import generate_pdf, generate_xlsx, generate_docx

router = APIRouter()

# mapiranje formata na generator + MIME tip + ekstenziju
FORMATS = {
    "PDF":  (generate_pdf,  "application/pdf", "pdf"),
    "XLSX": (generate_xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
    "DOCX": (generate_docx, "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "docx"),
}


@router.get("/meeting/{meeting_id}/export")
def export_report(
    meeting_id: int,
    format: str = "PDF",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Generiše izveštaj (zapisnik) za sastanak u traženom formatu.
    Format: PDF (default), XLSX ili DOCX.
    """
    fmt = format.upper()
    if fmt not in FORMATS:
        raise HTTPException(400, f"Nepodržan format. Dozvoljeni: {', '.join(FORMATS.keys())}")

    data = gather_meeting_data(meeting_id, db)
    if data is None:
        raise HTTPException(404, "Sastanak ne postoji")

    generator, mime, ext = FORMATS[fmt]
    content = generator(data)

    # evidentiraj generisanje izveštaja
    report = Report(
        meeting_id=meeting_id,
        requested_by=current_user.id,
        report_type="ZAPISNIK",
        period=fmt,
    )
    db.add(report)
    db.commit()

    filename = f"zapisnik_sastanak_{meeting_id}.{ext}"
    return StreamingResponse(
        io.BytesIO(content),
        media_type=mime,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/attendance-summary/me", response_model=AttendanceSummary)
def my_attendance_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Sumarni broj učešća prijavljenog korisnika (nedeljni/mesečni/godišnji)."""
    return _build_summary(current_user.id, db)


@router.get("/attendance-summary/team", response_model=list[AttendanceSummary])
def team_attendance_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Sumarni broj učešća za sve zaposlene u celini rukovodioca."""
    if not any(r in current_user._token_roles for r in ["RUKOVODILAC", "ADMIN"]):
        raise HTTPException(403, "Pristup zabranjen")
    # svi korisnici u istoj org celini
    team = db.query(User).filter(
        User.org_unit_id == current_user.org_unit_id,
        User.is_active == True,
    ).all()
    return [_build_summary(u.id, db) for u in team]


@router.get("/attendance-report")
def attendance_report(
    period: str = "MONTHLY",
    format: str = "PDF",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Izveštaj o učešću korisnika na sastancima za mesečni ili godišnji period.
    Sadrži temu, tačke dnevnog reda i datum održavanja.
    """
    now = datetime.now()
    if period.upper() == "MONTHLY":
        date_from = now.replace(day=1, hour=0, minute=0, second=0)
    elif period.upper() == "YEARLY":
        date_from = now.replace(month=1, day=1, hour=0, minute=0, second=0)
    else:
        raise HTTPException(400, "Period mora biti MONTHLY ili YEARLY")

    # sastanci na kojima je korisnik prisustvovao
    participations = db.query(MeetingParticipant).filter(
        MeetingParticipant.user_id == current_user.id,
        MeetingParticipant.attended == True,
    ).all()

    meeting_ids = [p.meeting_id for p in participations]
    meetings = db.query(Meeting).filter(
        Meeting.id.in_(meeting_ids),
        Meeting.scheduled_at >= date_from,
    ).order_by(Meeting.scheduled_at).all()

    # sastavi podatke za izveštaj
    items = []
    for m in meetings:
        agenda = db.query(AgendaItem).filter(
            AgendaItem.meeting_id == m.id
        ).order_by(AgendaItem.order_no).all()
        items.append({
            "topic": m.topic,
            "scheduled_at": m.scheduled_at,
            "status": m.status,
            "agenda_titles": [f"{a.order_no}. {a.title}" for a in agenda],
        })

    user_name = f"{current_user.first_name} {current_user.last_name}"
    period_label = "mesečni" if period.upper() == "MONTHLY" else "godišnji"

    fmt = format.upper()
    if fmt == "PDF":
        content = _generate_attendance_pdf(user_name, period_label, date_from, items)
        mime = "application/pdf"
        ext = "pdf"
    elif fmt == "XLSX":
        content = _generate_attendance_xlsx(user_name, period_label, date_from, items)
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    else:
        raise HTTPException(400, f"Nepodržan format. Dozvoljeni: PDF, XLSX")

    report = Report(
        requested_by=current_user.id,
        report_type="IZVESTAJ_UCESCA",
        period=period.upper(),
        meeting_id=None,
    )
    db.add(report)
    db.commit()

    period_ascii = "mesecni" if period.upper() == "MONTHLY" else "godisnji"
    filename = f"izvestaj_ucesca_{period_ascii}_{datetime.now().strftime('%Y_%m')}.{ext}"
    return StreamingResponse(
        io.BytesIO(content),
        media_type=mime,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_summary(user_id: int, db: Session) -> dict:
    """Broji učešća za korisnika: nedeljno, mesečno, godišnje."""
    now = datetime.now()
    week_start = now - timedelta(days=now.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0)
    month_start = now.replace(day=1, hour=0, minute=0, second=0)
    year_start = now.replace(month=1, day=1, hour=0, minute=0, second=0)

    user = db.query(User).filter(User.id == user_id).first()
    user_name = f"{user.first_name} {user.last_name}" if user else "N/A"

    base = db.query(MeetingParticipant).join(
        Meeting, MeetingParticipant.meeting_id == Meeting.id
    ).filter(
        MeetingParticipant.user_id == user_id,
        MeetingParticipant.attended == True,
    )

    weekly = base.filter(Meeting.scheduled_at >= week_start).count()
    monthly = base.filter(Meeting.scheduled_at >= month_start).count()
    yearly = base.filter(Meeting.scheduled_at >= year_start).count()

    return {
        "user_id": user_id,
        "user_name": user_name,
        "weekly": weekly,
        "monthly": monthly,
        "yearly": yearly,
    }


def _generate_attendance_pdf(user_name, period_label, date_from, items) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    font_dir = os.path.join(os.path.dirname(__import__("reportlab").__file__), "fonts")
    pdfmetrics.registerFont(TTFont("Vera", os.path.join(font_dir, "Vera.ttf")))
    pdfmetrics.registerFont(TTFont("Vera-Bold", os.path.join(font_dir, "VeraBd.ttf")))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    for s in styles.byName.values():
        s.fontName = "Vera"
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=16, fontName="Vera-Bold")
    story = []

    story.append(Paragraph(f"IZVEŠTAJ O UČEŠĆU — {period_label}", title_style))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"<b>Korisnik:</b> {user_name}", styles["Normal"]))
    story.append(Paragraph(f"<b>Period od:</b> {date_from.strftime('%d.%m.%Y')}", styles["Normal"]))
    story.append(Paragraph(f"<b>Broj sastanaka:</b> {len(items)}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))

    if items:
        tbl_data = [["Tema", "Datum", "Dnevni red"]]
        for it in items:
            tbl_data.append([
                it["topic"],
                it["scheduled_at"].strftime("%d.%m.%Y %H:%M"),
                "\n".join(it["agenda_titles"]) if it["agenda_titles"] else "-",
            ])
        t = Table(tbl_data, colWidths=[5*cm, 4*cm, 7*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Vera-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Vera"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Nema evidentiranih učešća u ovom periodu.", styles["Normal"]))

    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(
        f"<i>Generisano: {datetime.now().strftime('%d.%m.%Y %H:%M')}</i>",
        styles["Normal"],
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def _generate_attendance_xlsx(user_name, period_label, date_from, items) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Ucesca"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws["A1"] = f"IZVEŠTAJ O UČEŠĆU — {period_label}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    ws.cell(row=3, column=1, value="Korisnik:").font = bold
    ws.cell(row=3, column=2, value=user_name)
    ws.cell(row=4, column=1, value="Period od:").font = bold
    ws.cell(row=4, column=2, value=date_from.strftime("%d.%m.%Y"))
    ws.cell(row=5, column=1, value="Broj sastanaka:").font = bold
    ws.cell(row=5, column=2, value=len(items))

    r = 7
    for c, h in enumerate(["Tema", "Datum", "Dnevni red"], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    r = 8
    for it in items:
        ws.cell(row=r, column=1, value=it["topic"])
        ws.cell(row=r, column=2, value=it["scheduled_at"].strftime("%d.%m.%Y %H:%M"))
        ws.cell(row=r, column=3, value=", ".join(it["agenda_titles"]) if it["agenda_titles"] else "-")
        r += 1

    for col, w in zip("ABC", [30, 20, 40]):
        ws.column_dimensions[col].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@router.get("/", response_model=List[ReportOut])
def list_reports(
    db: Session = Depends(get_db),
    _=AnyAuthenticated,
):
    """Istorija generisanih izveštaja."""
    return db.query(Report).order_by(Report.created_at.desc()).all()


@router.get("/meeting/{meeting_id}", response_model=List[ReportOut])
def list_reports_for_meeting(
    meeting_id: int,
    db: Session = Depends(get_db),
    _=AnyAuthenticated,
):
    """Izveštaji generisani za konkretan sastanak."""
    return db.query(Report).filter(
        Report.meeting_id == meeting_id
    ).order_by(Report.created_at.desc()).all()